#!/usr/bin/env python

"""Aggregates evaluation results.
"""

import csv
import openpyxl
import pathlib
import os, re, sys


def explore_directory(dirpath):
    dirpath = pathlib.Path(dirpath)
    for filename in os.listdir(str(dirpath)):
        filepath = dirpath / filename
        if filepath.is_dir():
            for filepath in explore_directory(filepath):
                yield filepath
        else:
            yield filepath


headers = None
results = {}
fn_analysis_results = {}
timestamps = {}
runtimes = {}
runtime_str_pattern = re.compile(r'^([0-9]+):([0-9]+):([0-9]+)$')
for filepath in explore_directory('.'):
    if filepath.name == 'timings.csv':

        dataset_name = str(filepath.parents[len(filepath.parents) - 2])
        if not dataset_name in runtimes: runtimes[dataset_name] = {}
        dataset_runtimes = runtimes[dataset_name]

        total_column_idx = None
        with filepath.open('r', newline='') as csv_fin:
            csv_reader = csv.reader(csv_fin, delimiter=';', quotechar='|', quoting=csv.QUOTE_MINIMAL)
            for row in csv_reader:
                if total_column_idx is None and 'total' in row:
                    total_column_idx = row.index('total')
                elif row[0] == '':
                    assert total_column_idx is not None
                    task_path = os.sep.join(str(filepath.parent).split(os.sep)[1:])
                    runtime_str = row[total_column_idx]
                    runtime_match = runtime_str_pattern.match(runtime_str)
                    runtime = int(runtime_match.group(3)) + int(runtime_match.group(2)) * 60 + int(runtime_match.group(1)) * (60 ** 2)
                    dataset_runtimes[task_path] = runtime

    if filepath.name == 'study.csv':

        dataset_name = str(filepath.parents[len(filepath.parents) - 2])
        if not dataset_name in results: results[dataset_name] = {}
        dataset_results = results[dataset_name]

        with filepath.open('r', newline='') as csv_fin:
            csv_reader = csv.reader(csv_fin, delimiter=';', quotechar='|', quoting=csv.QUOTE_MINIMAL)
            for row in csv_reader:
                if row[0] == 'ID':
                    if headers is None:
                        headers = row[1:]
                        headers_filepath = filepath
                    elif headers != row[1:]:
                        print(f'Headers in {headers_filepath}: {", ".join(headers)}')
                        print(f'Headers in {filepath}: {", ".join(row[1:])}')
                        raise ValueError('incompatible headers')
                elif row[0] == '':
                    task_path = os.sep.join(str(filepath.parent).split(os.sep)[1:])
                    dataset_results[task_path] = row[1:]

        fn_results_filepath = filepath.parent / 'fn.csv'
        if fn_results_filepath.is_file():
            if dataset_name not in fn_analysis_results:
                fn_analysis_results[dataset_name] = {}
            fn_analysis_results[dataset_name][task_path] = fn_results_filepath
        
        if dataset_name not in timestamps:
            timestamps[dataset_name] = {}
        timestamps[dataset_name][task_path] = filepath.stat().st_mtime

def translate_header(header_name):
    return {
        'd/FN':    'FN',
        'd/FP':    'FP',
        'd/Merge': 'Merge',
        'd/Split': 'Split',
    }.get(header_name, header_name)

def write_xl_row(sheet, row, values):
    cells = []
    for column, value in enumerate(values, start=1):
        cell = sheet.cell(row=row, column=column)
        cell.value = value
        cells.append(cell)
    return cells

def is_baseline_task_path(task_path):
    return task_path.startswith('baseline')

def get_highscore_check(scores, task_paths, results, accept='all'):
    score_func = {
        'Dice':       max,
        'HSD (a2e)':  min,
        'HSD (e2a)':  min,
        'HSD* (a2e)': min,
        'HSD* (e2a)': min,
        'Jaccard':    max,
        'NSD':        min,
        'NSD*':       min,
        'Rand':       max,
        'SEG':        max,
        'FN':         min,
        'FP':         min,
        'Merge':      min,
        'Split':      min,
        'M+S':        min,
    }
    highscores = {}
    for task_path in task_paths:
        if (isinstance(accept, str) and accept == 'all') or accept(task_path):
            for score_type, score_value in zip(scores, results[task_path]):
                score_value = float(score_value)
                if score_type in highscores:
                    highscores[score_type] = score_func[score_type](highscores[score_type], score_value)
                else:
                    highscores[score_type] = score_value
    def create_highscore_check(score):
        def ret(value):
            if score in highscores:
                return score_func[score](value, highscores[score]) == value
            else:
                return False
        return ret
    return [create_highscore_check(score) for score in scores]

def get_baseline_check(scores, task_paths, results):
    return get_highscore_check(scores, task_paths, results, accept=\
        lambda task_path: is_baseline_task_path(task_path) and 'generic' in task_path)

if headers is not None:
    headers = [translate_header(header) for header in headers]
    for task_results in (tr for dr in results.values() for tr in dr.values()):
        false_merge = float(task_results[headers.index('Merge')])
        false_split = float(task_results[headers.index('Split')])
        task_results.append(str(false_merge + false_split))

    headers = headers + ['M+S', 'FN analysis', 'Runtime', 'Timestamp']
    default_font = openpyxl.styles.Font(size=10)
    header_font = openpyxl.styles.Font(size=10, bold=True)
    score_font = openpyxl.styles.Font(size=9, color='303030')
    highscore_font = openpyxl.styles.Font(size=9, bold=True)
    header_border = openpyxl.styles.borders.Border(bottom=openpyxl.styles.borders.Side(style='double'))
    task_path_border = openpyxl.styles.borders.Border(right=openpyxl.styles.borders.Side(style='thin'))
    last_baseline_border = openpyxl.styles.borders.Border(bottom=openpyxl.styles.borders.Side(style='thin'))
    fill_improved = openpyxl.styles.PatternFill(fgColor='AAFFAA', fill_type='solid')
    for dataset_name, dataset_results in results.items():
        task_paths = sorted(dataset_results.keys())
        is_improved = get_baseline_check(headers[:-1], task_paths, dataset_results)
        is_highscore = get_highscore_check(headers[:-1], task_paths, dataset_results, accept=lambda task_path: not (is_baseline_task_path(task_path) and 'generic' not in task_path))
        xlbook = openpyxl.Workbook()
        sheet  = xlbook.active
        cells = write_xl_row(sheet, 1, [dataset_name] + headers)
        for cell_idx, cell in enumerate(cells):
            cell.font = header_font
            cell.border = header_border
        cells[0].font = default_font
        for cell in cells[1:-3]:
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        sheet.column_dimensions['A'].width = max(len(task_path) for task_path in task_paths)
        for column, header in enumerate(headers, start=2):
            if column < len(headers) - 1: # score
                header_width = 10
            elif column == len(headers) - 1: # fn
                header_width = 12
            elif column == len(headers): # runtime
                header_width = 9
            else: # timestamp
                header_width = 18
            sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = header_width
        for row, task_path in enumerate(sorted(task_paths, key=lambda task_path: (0 if 'baseline' in task_path else 1, task_path)), start=2):
            try:
                task_results = dataset_results[task_path]
                cells = write_xl_row(sheet, row, [task_path] + [eval(val) for val in task_results])
                cells[0].font = header_font
                cells[0].border = task_path_border
                is_last_baseline_task_path = is_baseline_task_path(task_path) and not is_baseline_task_path(task_paths[row - 1])
                for cell_idx, cell in enumerate(cells[1:]):
                    if is_last_baseline_task_path:
                        cell.border = last_baseline_border
                    cell.font = score_font
                    cell.number_format = '#,##0.000'
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                    if is_highscore[cell_idx](float(task_results[cell_idx])):
                        cell.font = highscore_font
                    if not is_baseline_task_path(task_path) and is_improved[cell_idx](float(task_results[cell_idx])):
                        cell.fill = fill_improved
                fn_analysis_cell = sheet.cell(row=row, column=len(task_results) + 2)
                fn_analysis_cell.font = default_font
                runtime_cell = sheet.cell(row=row, column=len(task_results) + 3)
                if is_last_baseline_task_path:
                    fn_analysis_cell.border = last_baseline_border
                    runtime_cell.border = last_baseline_border
                if dataset_name in fn_analysis_results and task_path in fn_analysis_results[dataset_name]:
                    fn_analysis_filepath = fn_analysis_results[dataset_name][task_path]
                    fn_analysis_cell.value = fn_analysis_filepath.name
                    fn_analysis_cell.hyperlink = f'{str(fn_analysis_filepath)}'
                if task_path in runtimes[dataset_name]:
                    runtime_cell.font = default_font
                    runtime_cell.number_format = 'HH:MM:SS'
                    runtime_cell.value = f'={runtimes[dataset_name][task_path]} / (60 * 60 * 24)'
                    runtime_cell.alignment = openpyxl.styles.Alignment(horizontal='left')
                timestamp_cell = sheet.cell(row=row, column=len(task_results) + 4)
                if is_last_baseline_task_path:
                    timestamp_cell.border = last_baseline_border
                elif not is_baseline_task_path(task_path):
                    timestamp_cell.font = default_font
                    timestamp_cell.number_format = 'MMM D, YYYY, H:MM'
                    timestamp_cell.value = f'={timestamps[dataset_name][task_path] / 86400} + DATE(1970,1,1)'
                    timestamp_cell.alignment = openpyxl.styles.Alignment(horizontal='left')
            except:
                print(f'Error while processing task: {dataset_name}/{task_path} ({repr(sys.exc_info()[1])})')
                raise
        xlbook.save(f'{dataset_name}.xlsx')

